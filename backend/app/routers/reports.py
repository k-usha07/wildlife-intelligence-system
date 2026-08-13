from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from io import BytesIO

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.survey import Survey
from app.models.observation import Observation
from app.models.biodiversity import BiodiversityIndex

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get("/survey/{survey_id}/pdf")
async def export_survey_pdf(
    survey_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors

    survey_result = await db.execute(select(Survey).where(Survey.id == survey_id))
    survey = survey_result.scalar_one_or_none()
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")

    obs_result = await db.execute(select(Observation).where(Observation.survey_id == survey_id))
    observations = obs_result.scalars().all()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Wildlife Survey Report", styles["Title"]))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Survey: {survey.name}", styles["Heading2"]))
    elements.append(Paragraph(f"Status: {survey.status}", styles["Normal"]))
    elements.append(Paragraph(f"Start Date: {survey.start_date}", styles["Normal"]))
    if survey.end_date:
        elements.append(Paragraph(f"End Date: {survey.end_date}", styles["Normal"]))
    elements.append(Paragraph(f"Total Observations: {len(observations)}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    if observations:
        elements.append(Paragraph("Observation Details", styles["Heading2"]))
        elements.append(Spacer(1, 10))
        headers = ["Species", "Count", "Confidence", "Endangered"]
        table_data = [headers]
        for obs in observations[:50]:
            table_data.append([
                obs.species_name or "N/A",
                str(obs.count_estimate or 1),
                f"{obs.confidence:.2f}" if obs.confidence else "N/A",
                "Yes" if obs.is_endangered else "No",
            ])
        table = Table(table_data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=survey_{survey_id}_report.pdf"},
    )


@router.get("/survey/{survey_id}/excel")
async def export_survey_excel(
    survey_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font

    obs_result = await db.execute(select(Observation).where(Observation.survey_id == survey_id))
    observations = obs_result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Observations"

    headers = ["ID", "Species", "Count", "Confidence", "Endangered", "Behavior", "Verified"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row, obs in enumerate(observations, 2):
        ws.cell(row=row, column=1, value=str(obs.id))
        ws.cell(row=row, column=2, value=obs.species_name or "N/A")
        ws.cell(row=row, column=3, value=obs.count_estimate or 1)
        ws.cell(row=row, column=4, value=obs.confidence)
        ws.cell(row=row, column=5, value="Yes" if obs.is_endangered else "No")
        ws.cell(row=row, column=6, value=obs.behavior or "")
        ws.cell(row=row, column=7, value="Yes" if obs.is_verified else "No")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=survey_{survey_id}_report.xlsx"},
    )


@router.get("/biodiversity/{survey_id}")
async def biodiversity_report(
    survey_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    bio_result = await db.execute(
        select(BiodiversityIndex).where(BiodiversityIndex.survey_id == survey_id)
        .order_by(BiodiversityIndex.calculation_date.desc())
    )
    bio_data = bio_result.scalars().all()

    return {
        "survey_id": str(survey_id),
        "biodiversity_records": [
            {
                "shannon_index": b.shannon_index,
                "simpson_index": b.simpson_index,
                "species_richness": b.species_richness,
                "evenness": b.evenness_index,
                "area_km2": b.area_km2,
                "date": b.calculation_date.isoformat() if b.calculation_date else None,
            }
            for b in bio_data
        ],
    }